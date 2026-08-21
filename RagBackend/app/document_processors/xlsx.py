"""
XLSX extraction: format-preserving, multi-row headers, wide/large sheet handling.
"""

from typing import List, Any
import pandas as pd
from langchain_core.documents import Document

from app.config import get_logger

logger = get_logger("askdoc.processors.xlsx")


def extract_xlsx_sheets(xlsx_path: str, filename: str) -> List[Document]:
    """
    Advanced structure-preserving XLSX parser:
    1. Preserves cell display formatting (currencies, percentages, thousands separators).
    2. Detects multi-row headers (e.g. Row 1 Category + Row 2 Metric) and combines them into compound headers.
    3. Handles wide sheets (> 15 columns) by partitioning or noting truncation.
    4. Handles large sheets (> 500 rows) by batching 5-10 rows per chunk.
    5. Prepends sheet_name and column context to all chunks.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not available, falling back to pandas excel parser")
        excel_sheets = pd.read_excel(xlsx_path, sheet_name=None)
        documents = []
        for sheet_name, df in excel_sheets.items():
            if df.empty:
                continue
            col_names = ", ".join([str(c) for c in df.columns])
            for idx, row in df.iterrows():
                row_str = "\n".join([f"{col}: {val}" for col, val in row.items() if pd.notna(val)])
                content = f"File: {filename} | Sheet: {sheet_name}\nColumns: {col_names}\nRow Data:\n{row_str}"
                documents.append(Document(page_content=content, metadata={"source": filename, "sheet": sheet_name, "row": idx}))
        return documents

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        logger.error(f"Error opening XLSX file {filename}: {e}", exc_info=True)
        return []

    documents = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows = []
        for row in ws.iter_rows(values_only=False):
            row_data = []
            for cell in row:
                val = cell.value
                if val is None:
                    row_data.append("")
                    continue

                fmt = cell.number_format or ""
                # Format preservation: Currency, Percentage, Integers, Decimals
                if isinstance(val, (int, float)):
                    if "%" in fmt:
                        row_data.append(f"{val * 100:.1f}%" if abs(val) < 10 else f"{val}%")
                    elif any(curr in fmt for curr in ["$", "₹", "€", "£", "¥"]):
                        symbol = next((curr for curr in ["$", "₹", "€", "£", "¥"] if curr in fmt), "$")
                        row_data.append(f"{symbol}{val:,.2f}" if isinstance(val, float) else f"{symbol}{val:,}")
                    elif isinstance(val, float) and val.is_integer():
                        row_data.append(str(int(val)))
                    elif isinstance(val, float):
                        row_data.append(f"{val:,.2f}" if abs(val) >= 1000 else f"{val:.4g}")
                    else:
                        row_data.append(str(val))
                else:
                    row_data.append(str(val).strip())

            if any(c != "" for c in row_data):
                raw_rows.append(row_data)

        if not raw_rows:
            continue

        # 1. Multi-row header detection (e.g. Row 0 Category + Row 1 Subheader in financial sheets)
        is_multi_row_header = False
        if len(raw_rows) >= 3:
            row0_has_blanks = any(c == "" for c in raw_rows[0]) and any(c != "" for c in raw_rows[0])
            row0_non_empty = [c for c in raw_rows[0] if c]
            row1_non_empty = [c for c in raw_rows[1] if c]
            if row0_has_blanks and len(row1_non_empty) >= 2:
                non_num_count = sum(1 for c in row1_non_empty if not c.replace(",", "").replace(".", "").replace("$", "").replace("%", "").isdigit())
                if (non_num_count / len(row1_non_empty)) >= 0.7:
                    is_multi_row_header = True

        if is_multi_row_header:
            # Forward-fill merged category headers from row 0
            filled_row0 = []
            last_cat = ""
            for c in raw_rows[0]:
                if c:
                    last_cat = c
                filled_row0.append(last_cat)

            combined_headers = []
            for col_idx, (cat, sub) in enumerate(zip(filled_row0, raw_rows[1])):
                if cat and sub and cat != sub:
                    combined_headers.append(f"{cat} - {sub}")
                else:
                    combined_headers.append(sub or cat or f"Col_{col_idx+1}")
            data_rows = raw_rows[2:]
        else:
            combined_headers = [c if c else f"Col_{i+1}" for i, c in enumerate(raw_rows[0])]
            data_rows = raw_rows[1:]

        if not data_rows:
            continue

        # 2. Wide Sheet handling (> 15 columns)
        is_truncated = False
        col_limit = 15
        active_headers = combined_headers
        if len(combined_headers) > col_limit:
            is_truncated = True
            active_headers = combined_headers[:col_limit]

        col_names = ", ".join(active_headers)
        total_rows = len(data_rows)

        # 3. Large Sheet Batching (> 500 rows)
        if total_rows > 500:
            batch_size = 8
            for start_idx in range(0, total_rows, batch_size):
                batch_end = min(start_idx + batch_size, total_rows)
                batch_rows = data_rows[start_idx:batch_end]
                
                rows_text = []
                for row_offset, row in enumerate(batch_rows):
                    row_num = start_idx + row_offset + 1
                    row_cells = row[:len(active_headers)] if is_truncated else row
                    row_pairs = [f"{col}: {val}" for col, val in zip(active_headers, row_cells) if val != ""]
                    if row_pairs:
                        rows_text.append(f"[Row {row_num}] " + "; ".join(row_pairs))

                if rows_text:
                    content = (
                        f"File: {filename} | Sheet: {sheet_name}\n"
                        f"Columns: {col_names}\n"
                        f"Rows {start_idx + 1}-{batch_end} Data:\n" +
                        "\n".join(rows_text)
                    )
                    documents.append(Document(
                        page_content=content,
                        metadata={
                            "source": filename,
                            "title": filename,
                            "sheet": sheet_name,
                            "sheet_name": sheet_name,
                            "row_start": start_idx + 1,
                            "row_end": batch_end,
                            "row": f"{start_idx + 1}-{batch_end}",
                            "total_rows": total_rows,
                            "is_truncated": is_truncated
                        }
                    ))
        else:
            # Single-row precision for sheets <= 500 rows
            for row_idx, row in enumerate(data_rows):
                row_num = row_idx + 1
                row_cells = row[:len(active_headers)] if is_truncated else row
                row_pairs = [f"{col}: {val}" for col, val in zip(active_headers, row_cells) if val != ""]
                if row_pairs:
                    row_str = "\n".join(row_pairs)
                    content = f"File: {filename} | Sheet: {sheet_name}\nColumns: {col_names}\nRow Data:\n{row_str}"
                    documents.append(Document(
                        page_content=content,
                        metadata={
                            "source": filename,
                            "title": filename,
                            "sheet": sheet_name,
                            "sheet_name": sheet_name,
                            "row": row_num,
                            "total_rows": total_rows,
                            "is_truncated": is_truncated
                        }
                    ))

    return documents
